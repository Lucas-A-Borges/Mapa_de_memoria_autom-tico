import pandas as pd
import xml.etree.ElementTree as ET
import os
import shutil
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Font, Side, Border
from datetime import datetime
from typing import Dict, Any, List
import sys



# --- CONFIGURAÇÕES GLOBAIS ---
SUFIXOS = [
    'EST1','EST2', 'DEF1','DEF2','DEF3','DEF4', 
    'INT1','INT2','INT3', 'INT4', 'CMD1','CMD2', 
    'ALR1', 'ALR2', 'ALR3'
]
ARQUIVO_UNITPRO = 'unitpro.xef'
TIPOS_PERMITIDOS = ['WORD', 'BOOL', 'EBOOL']
TIPOS_PARA_VARIAVEL_PRINCIPAL = ['WORD'] 
SUFIXOS_DEFEITO = ['DEF1', 'DEF2', 'DEF3', 'DEF4']
SUFIXOS_CMD_EST = ['CMD1', 'CMD2', 'EST1', 'EST2']

Sobrescrever_set_bit = True # Defina como False para que SET_BIT não sobrescreva TAGs já preenchidas 
Habilitar_ordenacao = True  # Defina como False para desabilitar a ordenação final
# ----------------------------------------------------------------------
# CONSTANTES POR TIPO DE BLOCO (Tags Fixas de Prioridade 1)
# ----------------------------------------------------------------------

MOTOR_BITS_INFO = {
    "DEF": {
        0: {"tag_suffix": "_DSECC", "comment": "Seccionadora aberta"},
        1: {"tag_suffix": "_DSOBR", "comment": "Térmico atuado"},
        2: {"tag_suffix": "_DPART", "comment": "Defeito de partida"},
        3: {"tag_suffix": "_DDESL", "comment": "Botão desliga atuado"},
        4: {"tag_suffix": "_DCOM",  "comment": "Falha na comunicação"},
        5: {"tag_suffix": "_DCONT", "comment": "Defeito no contator"},
        6: {"tag_suffix": "_DEMER", "comment": "Emergência atuada"},
        7: {"tag_suffix": "_DSAL",  "comment": "Velocidade baixa"},
        8: {"tag_suffix": "_DDEF",  "comment": "Defeito elétrico"}
    },
    "CMD": {
        8:  {"tag_suffix": "_CLIGAR", "comment": "Comando liga reverso"},
        9:  {"tag_suffix": "_CREAR",  "comment": "Comando rearme"},
        10: {"tag_suffix": "_CMAN",   "comment": "Comando manual"},
        11: {"tag_suffix": "_CAUTO",  "comment": "Comando Automático"},
        12: {"tag_suffix": "_CDESL",  "comment": "Comando desliga"},
        13: {"tag_suffix": "_CLIGAF", "comment": "Comando liga frente"},
        14: {"tag_suffix": "_CLOC",   "comment": "Comando Local"},
        15: {"tag_suffix": "_CREM",   "comment": "Comando Remoto"},
    },
    "EST": {
        1:  {"tag_suffix": "_EFUNF",  "comment": "Estado funcionando frente"},
        2:  {"tag_suffix": "_EFUNR",  "comment": "Estado funcionando ré"},
        7:  {"tag_suffix": "_EPART",  "comment": "Estado partindo"},
        10: {"tag_suffix": "_EMA",    "comment": "Estado manual automático"},
        11: {"tag_suffix": "_ELR",    "comment": "Estado local remoto"},
        12: {"tag_suffix": "_EFUNC",  "comment": "Estado funcionando"},
        13: {"tag_suffix": "_EPANT",  "comment": "Permissão anterior"},
        15: {"tag_suffix": "_EDEF",   "comment": "Estado de defeito"},
    }
}

VAL_BITS_INFO = {
    "DEF": {
        0: {"tag_suffix": "_DARAB", "comment": "Defeito arrancada abertura"},
        1: {"tag_suffix": "_DABER", "comment": "Defeito abertura"},
        2: {"tag_suffix": "_DAFEC", "comment": "Defeito arrancada fechamento"},
        3: {"tag_suffix": "_DFECH", "comment": "Defeito fechamento"}
    },
    "CMD": {
        6:  {"tag_suffix": "_CFECH", "comment": "Comando fecha válvula"},
        7:  {"tag_suffix": "_CABRE", "comment": "Comando abre válvula"},
        9:  {"tag_suffix": "_CREAR", "comment": "Comando rearme"},
        10: {"tag_suffix": "_CMAN",  "comment": "Comando manual"},    
        11: {"tag_suffix": "_CAUTO", "comment": "Comando Automático"},
        14: {"tag_suffix": "_CLOC",  "comment": "Comando Local"},     
        15: {"tag_suffix": "_CREM",  "comment": "Comando Remoto"},
    },
    "EST": {
        8:  {"tag_suffix": "_EZALF",  "comment": "Estado fechado da válvula"},
        9:  {"tag_suffix": "_EZAHA",  "comment": "Estado aberto da válvula"},
        10: {"tag_suffix": "_EMA",    "comment": "Estado manual automático"},
        11: {"tag_suffix": "_ELR",    "comment": "Estado local remoto"},
        12: {"tag_suffix": "_EFUNC",  "comment": "Indicação de funcionando para abrir ou fechar"},
        13: {"tag_suffix": "_EPANT",  "comment": "Indicação de permissão para abrir ou fechar"},
        15: {"tag_suffix": "_EDEF",   "comment": "Estado de defeito"},
    }
}

# ----------------------------------------------------------------------
# CONSTANTES PARA BLOCO SEQ (Prioridade 4)
# ----------------------------------------------------------------------
# 1. CMD1 (Tags Fixas - Usadas pela função padronizada aplicar_regras_fixas_bloco)
SEQ_CMD1_INFO = {
    9:  {"tag_suffix": "_CREAR", "comment": "Comando rearme"},
    10: {"tag_suffix": "_CMAN",  "comment": "Comando manual"},    
    11: {"tag_suffix": "_CAUTO", "comment": "Comando Automático"},
    12: {"tag_suffix": "_CDESL", "comment": "Comando desliga"},
    13: {"tag_suffix": "_CLIGA", "comment": "Comando liga"},
    14: {"tag_suffix": "_CLOC",  "comment": "Comando Local"},
    15: {"tag_suffix": "_CREM",  "comment": "Comando Remoto"},
}

# 2. DEF1 (Tags Fixas - Usadas pela função padronizada aplicar_regras_fixas_bloco)
SEQ_DEF1_INFO = {
    0:  {"tag_suffix": "_DTMAX1", "comment": "Tempo máximo de partida"},
    1:  {"tag_suffix": "_DTMAX2", "comment": "Tempo máximo de parada"},
}

# 3. EST1 (Misto: Saída Formal + Tags Fixas - Exige iteração manual)
SEQ_EST1_MAPPING = {
    4:  {"type": "FORMAL_OUT", "param": "EPART", "comment": "Estado partindo"},
    5:  {"type": "FORMAL_OUT", "param": "EPARA", "comment": "Estado parando"},
    10: {"type": "TAG_FIXA",   "suffix": "_EMA",  "comment": "Estado manual automático"}, 
    11: {"type": "TAG_FIXA",   "suffix": "_ELR",  "comment": "Estado local remoto"},     
    12: {"type": "FORMAL_OUT", "param": "EFUNC", "comment": "Estado funcionando"},
    13: {"type": "TAG_FIXA",   "suffix": "_EPANT", "comment": "Permissão anterior"}, 
    15: {"type": "FORMAL_OUT", "param": "EDEF",  "comment": "Estado de defeito"},
}


# ----------------------------------------------------------------------
# FUNÇÕES DE UTILIDADE
# ----------------------------------------------------------------------

def ler_variaveis_unitpro(caminho_arquivo: str) -> List[Dict[str, str]]:
    """Lê todas as variáveis do unitpro.xef."""
    # ... (Implementação omitida por brevidade, assumida como funcional)
    lista_variaveis = []
    try:
        tree = ET.parse(caminho_arquivo)
        root = tree.getroot()
    except (FileNotFoundError, ET.ParseError) as e:
        print(f"ERRO: Não foi possível ler ou fazer o parse do arquivo: {e}")
        return lista_variaveis

    for var_element in root.findall('.//variables'):
        nome = var_element.get('name')
        tipo = var_element.get('typeName')
        endereco = var_element.get('topologicalAddress')
        comentario_element = var_element.find('comment')
        comentario = comentario_element.text.strip() if comentario_element is not None and comentario_element.text else ""

    

        if nome and tipo in TIPOS_PERMITIDOS: # and endereco
            lista_variaveis.append({
                'nome': nome,
                'comentario': comentario,
                'endereco': endereco,
                'tipo': tipo
            })
    return lista_variaveis


def criar_estrutura_equipamento(nome_equipamento: str) -> Dict[str, Any]:
    """Cria a estrutura de dados inicial para um equipamento."""
    equipamento_estrutura = {
        'nome': nome_equipamento,
        'sufixos': {}
    }
    for sufixo in SUFIXOS:
        variaveis_bit = [{'nome': '', 'comentario': ''} for i in range(16)]
        equipamento_estrutura['sufixos'][sufixo] = {
            'nome': '',
            'endereco': '',
            'variaveis': variaveis_bit,
            # 'controle_total': False  <-- REMOVIDO
        }
    return equipamento_estrutura


def catalogar_variaveis(lista_variaveis: List[Dict[str, str]]) -> Dict[str, Dict[str, Any]]:
    """Organiza as variáveis lidas na matriz de equipamentos."""
    # ... (Implementação omitida por brevidade, assumida como funcional)
    matriz_equipamentos: Dict[str, Dict[str, Any]] = {}
    
    for var in lista_variaveis:
        if var['tipo'] not in TIPOS_PARA_VARIAVEL_PRINCIPAL: 
            continue
            
        for sufixo in SUFIXOS:
            parte_sufixo = f'_{sufixo}'
            if var['nome'].endswith(parte_sufixo): 
                nome_equipamento = var['nome'][:-len(parte_sufixo)]
                
                if nome_equipamento not in matriz_equipamentos:
                    matriz_equipamentos[nome_equipamento] = criar_estrutura_equipamento(nome_equipamento)
                    
                matriz_equipamentos[nome_equipamento]['sufixos'][sufixo].update({
                    'nome': var['nome'],
                    'endereco': var['endereco']
                })
                break
                
    return matriz_equipamentos


def aplicar_regras_fixas_bloco(equipamento_nome: str, sufixo_destino: str, dados_sufixo: Dict[str, Any], regras: Dict[int, Dict[str, str]], sobrescrever: bool = True):
    """Função auxiliar padronizada para aplicar regras fixas de CMD/EST/DEF."""
    variaveis_bit = dados_sufixo['variaveis']
    
    for bit, info in regras.items():
        tag_completa = f"{equipamento_nome}{info['tag_suffix']}"
        
        # Lógica de Sobrescrita:
        # Se 'sobrescrever' é True (Prioridade 1), ou se o slot estiver vazio (Prioridade 2+)
        if sobrescrever or not variaveis_bit[bit]['nome']:
            variaveis_bit[bit]['nome'] = tag_completa
            variaveis_bit[bit]['comentario'] = info['comment']


# ----------------------------------------------------------------------
# FUNÇÕES DE LÓGICA DE PREENCHIMENTO (Prioridade 1 - Sobrescreve)
# ----------------------------------------------------------------------

def preencher_bits_bloco_motor(caminho_arquivo: str, matriz_equipamentos: Dict[str, Dict[str, Any]]) -> Dict[str, Dict[str, Any]]:
    """Prioridade 1 - Preenche DEF, CMD e EST (tags fixas) - SOBRESCREVE."""
    try:
        tree = ET.parse(caminho_arquivo)
        root = tree.getroot()
    except (FileNotFoundError, ET.ParseError):
        return matriz_equipamentos

    for ffb_block in root.findall('.//FFBBlock'):
        block_type = ffb_block.get('typeName')

        if not block_type or not block_type.startswith('MOT'):
            continue

        desc_ffb = ffb_block.find('descriptionFFB')
        if desc_ffb is None: continue

        # 1. ENCONTRAR O EQUIPAMENTO (Usando a saída formal 'DEF' que deve ser EQ_DEF1)
        pin_def = desc_ffb.find('./outputVariable[@formalParameter="DEF1"]')
        target_name_def = pin_def.get('effectiveParameter') if pin_def is not None else None
        
        # Corrigido: Verifica se o pino DEF está conectado a uma palavra DEF1
        if not target_name_def or not target_name_def.endswith('_DEF1'):
            continue

        equipamento = target_name_def[:-len('_DEF1')]
        
        if equipamento not in matriz_equipamentos:
            continue

        # 2. PREPARAR DESTINOS DE PREENCHIMENTO (Garantindo que DEF1, CMD1 e EST1 estejam presentes no projeto)
        destinos_a_preencher = {}
        for sufixo_controle in ['CMD1', 'EST1', 'DEF1']:
            dados_sufixo = matriz_equipamentos[equipamento]['sufixos'][sufixo_controle]
            
            if dados_sufixo['nome']: 
                destinos_a_preencher[sufixo_controle] = dados_sufixo
        
        if not destinos_a_preencher: 
             continue

        # 3. COLETAR DF_BITS (bits 9 a 15 de DEF do XML)
        df_bits = {} 
        for i in range(9, 16):
            formal_param = f'DF_BIT{i:02}'
            input_variable = desc_ffb.find(f'./inputVariable[@formalParameter="{formal_param}"]')
            effective_param = input_variable.get('effectiveParameter') if input_variable is not None else None
            df_bits[i] = effective_param if effective_param and effective_param != '0' else ''

        # 4. APLICAR A REGRA (SOBRESCREVE)
        for sufixo, dados_sufixo in destinos_a_preencher.items():
            # Flag 'controle_total' removida

            if sufixo.startswith('DEF'):
                # a. Preencher DEF bits 0 a 8 (Fixos)
                aplicar_regras_fixas_bloco(equipamento, sufixo, dados_sufixo, MOTOR_BITS_INFO["DEF"], sobrescrever=True)
                
                # b. Preencher DEF bits 9 a 15 (DF_BITs do XML)
                variaveis_bit = dados_sufixo['variaveis']
                for i in range(9, 16):
                    if df_bits[i]:
                        variaveis_bit[i]['nome'] = df_bits[i]
                        #variaveis_bit[i]['comentario'] = 'Defeito Auxiliar FFB'
                        
            elif sufixo.startswith('CMD'):
                aplicar_regras_fixas_bloco(equipamento, sufixo, dados_sufixo, MOTOR_BITS_INFO["CMD"], sobrescrever=True)
                
            elif sufixo.startswith('EST'):
                aplicar_regras_fixas_bloco(equipamento, sufixo, dados_sufixo, MOTOR_BITS_INFO["EST"], sobrescrever=True)

    return matriz_equipamentos


def preencher_bits_bloco_valvula(caminho_arquivo: str, matriz_equipamentos: Dict[str, Dict[str, Any]]) -> Dict[str, Dict[str, Any]]:
    """Prioridade 1 - Preenche DEF, CMD e EST (tags fixas) - SOBRESCREVE."""
    try:
        tree = ET.parse(caminho_arquivo)
        root = tree.getroot()
    except (FileNotFoundError, ET.ParseError) as e:
        return matriz_equipamentos

    #Procura blocos do tipo VALVULA
    for ffb_block in root.findall('.//FFBBlock'):
        block_type = ffb_block.get('typeName')
        if not block_type or not (block_type.startswith('VAL') or block_type.startswith('ABRI')):
            continue
        
        desc_ffb = ffb_block.find('descriptionFFB')
        if desc_ffb is None: continue

        # 1. ENCONTRAR O EQUIPAMENTO (Usando a saída formal 'DEF' ou 'EST')
        equipamento = None
        for formal_param in ["DEF", "EST"]:
             out_variable = desc_ffb.find(f'./outputVariable[@formalParameter="{formal_param}"]')
             if out_variable is not None:
                target_name = out_variable.get('effectiveParameter')
                if target_name and target_name.endswith('_DEF1'):
                    equipamento = target_name[:-len('_DEF1')]
                    break
                elif target_name and target_name.endswith('_EST1'):
                    equipamento = target_name[:-len('_EST1')]
                    break
        
        if not equipamento or equipamento not in matriz_equipamentos:
            continue

        # 2. PREPARAR DESTINOS DE PREENCHIMENTO (Garantindo que DEF1, CMD1 e EST1 existam)
        destinos_a_preencher = {}
        for sufixo_controle in ['CMD1', 'EST1', 'DEF1']:
            dados_sufixo = matriz_equipamentos[equipamento]['sufixos'][sufixo_controle]
            
            if dados_sufixo['nome']: 
                destinos_a_preencher[sufixo_controle] = dados_sufixo
        
        if not destinos_a_preencher: 
             continue

        # 3. APLICAR A REGRA (SOBRESCREVE)
        for sufixo, dados_sufixo in destinos_a_preencher.items():
            # Flag 'controle_total' removida

            if sufixo.startswith('DEF'):
                aplicar_regras_fixas_bloco(equipamento, sufixo, dados_sufixo, VAL_BITS_INFO["DEF"], sobrescrever=True)
            elif sufixo.startswith('CMD'):
                aplicar_regras_fixas_bloco(equipamento, sufixo, dados_sufixo, VAL_BITS_INFO["CMD"], sobrescrever=True)
            elif sufixo.startswith('EST'):
                aplicar_regras_fixas_bloco(equipamento, sufixo, dados_sufixo, VAL_BITS_INFO["EST"], sobrescrever=True)

    return matriz_equipamentos


# ----------------------------------------------------------------------
# FUNÇÕES DE LÓGICA DE PREENCHIMENTO (Prioridade 2+ - Não sobrescreve)
# ----------------------------------------------------------------------

# ... (preencher_bits_bit_to_word e preencher_bits_set_bit permanecem inalteradas, usando sobrescrever=False)
def preencher_bits_bit_to_word(caminho_arquivo: str, matriz_equipamentos: Dict[str, Dict[str, Any]]) -> Dict[str, Dict[str, Any]]:
    """Prioridade 2 - Preenche baseado em conexões XML - Não sobrescreve."""
    # ... (Lógica mantida, garantindo que não sobrescreva)
    try:
        tree = ET.parse(caminho_arquivo)
        root = tree.getroot()
    except (FileNotFoundError, ET.ParseError) as e:
        return matriz_equipamentos

    for ffb_block in root.findall('.//FFBBlock[@typeName="BIT_TO_WORD"]'):
        desc_ffb = ffb_block.find('descriptionFFB')
        if desc_ffb is None: continue
        out_variable = desc_ffb.find('./outputVariable[@formalParameter="OUT"]')
        target_name = out_variable.get('effectiveParameter') if out_variable is not None else None
        if not target_name: continue

        equipamento = None
        sufixo_destino = None

        for sufixo in SUFIXOS:
            parte_sufixo = f'_{sufixo}'
            if target_name.endswith(parte_sufixo):
                nome_equipamento = target_name[:-len(parte_sufixo)]
                if nome_equipamento in matriz_equipamentos:
                    equipamento = nome_equipamento
                    sufixo_destino = sufixo
                    break
        
        if not equipamento or not sufixo_destino: continue


        variaveis_bit = matriz_equipamentos[equipamento]['sufixos'][sufixo_destino]['variaveis']
        
        for i in range(16):
            formal_param = f'BIT{i}'
            input_variable = desc_ffb.find(f'./inputVariable[@formalParameter="{formal_param}"]')
            
            if input_variable is not None:
                effective_param = input_variable.get('effectiveParameter')
                if effective_param:
                    # Não sobrescrever preenchimento existente
                    if not variaveis_bit[i]['nome']: 
                        variaveis_bit[i]['nome'] = effective_param
            
    return matriz_equipamentos

def preencher_bits_set_bit(caminho_arquivo: str, matriz_equipamentos: Dict[str, Dict[str, Any]]) -> Dict[str, Dict[str, Any]]:
    """Prioridade 3 - Preenche baseado em conexões XML - Não sobrescreve."""
    # ... (Lógica mantida, garantindo que não sobrescreva)
    try:
        tree = ET.parse(caminho_arquivo)
        root = tree.getroot()
    except (FileNotFoundError, ET.ParseError) as e:
        return matriz_equipamentos

    for ffb_block in root.findall('.//FFBBlock[@typeName="SET_BIT"]'):
        desc_ffb = ffb_block.find('descriptionFFB')
        if desc_ffb is None: continue

        out_variable = desc_ffb.find('./outputVariable[@formalParameter="RES"]')
        target_name = out_variable.get('effectiveParameter') if out_variable is not None else None
        if not target_name: continue

        in_variable = desc_ffb.find('./inputVariable[@formalParameter="IN"]')
        effective_param_in = in_variable.get('effectiveParameter') if in_variable is not None else None
        if not effective_param_in: continue
        
        no_variable = desc_ffb.find('./inputVariable[@formalParameter="NO"]')
        bit_no_str = no_variable.get('effectiveParameter') if no_variable is not None else None
        
        try:
            bit_index = int(bit_no_str) - 1 
            if not (0 <= bit_index <= 15): continue
        except (ValueError, TypeError):
             continue

        equipamento = None
        sufixo_destino = None

        for sufixo in SUFIXOS:
            parte_sufixo = f'_{sufixo}'
            if target_name.endswith(parte_sufixo):
                nome_equipamento = target_name[:-len(parte_sufixo)]
                if nome_equipamento in matriz_equipamentos:
                    equipamento = nome_equipamento
                    sufixo_destino = sufixo
                    break
        


        if not equipamento or not sufixo_destino: continue

        variaveis_bit = matriz_equipamentos[equipamento]['sufixos'][sufixo_destino]['variaveis']
        
        # Não sobrescrever preenchimento existente
        
        if Sobrescrever_set_bit or not variaveis_bit[bit_index]['nome']: # Não sobrescrever preenchimento existente
            variaveis_bit[bit_index]['nome'] = effective_param_in
            
    return matriz_equipamentos

def preencher_bits_bloco_seq(caminho_arquivo: str, matriz_equipamentos: Dict[str, Dict[str, Any]]) -> Dict[str, Dict[str, Any]]:
    """Prioridade 4 - Preenche EST1 (Misto), DEF1 e CMD1 (Tags Fixas Padronizadas) - Não sobrescreve."""
    try:
        tree = ET.parse(caminho_arquivo)
        root = tree.getroot()
    except (FileNotFoundError, ET.ParseError):
        return matriz_equipamentos

    for ffb_block in root.findall('.//FFBBlock'):
        block_type = ffb_block.get('typeName')
        if not block_type or not block_type.startswith('SEQ'):
            continue

        desc_ffb = ffb_block.find('descriptionFFB')
        if desc_ffb is None: continue

        # 1. Identificar o EQUIPAMENTO (Usando a saída EST1 ou DEF1)
        equipamento = None
        for sufixo_check in ['DEF1', 'EST1']:
             out_variable = desc_ffb.find(f'./outputVariable[@formalParameter="{sufixo_check}"]')
             if out_variable is not None:
                target_name = out_variable.get('effectiveParameter')
                if target_name and target_name.endswith(f'_{sufixo_check}'):
                    equipamento = target_name[:-len(f'_{sufixo_check}')]
                    break
        
        if not equipamento or equipamento not in matriz_equipamentos:
            continue
            
        eq_data = matriz_equipamentos[equipamento]
        
        # 2. Preencher CMD1 e DEF1 usando o método padronizado (Aplicar Regras Fixas)
        
        # CMD1 (Tags Fixas - Prioridade 4, Não Sobrescreve)
        dados_cmd1 = eq_data['sufixos']['CMD1']
        if dados_cmd1['nome']:
            aplicar_regras_fixas_bloco(equipamento, 'CMD1', dados_cmd1, SEQ_CMD1_INFO, sobrescrever=False)

        # DEF1 (Tags Fixas - Prioridade 4, Não Sobrescreve)
        dados_def1 = eq_data['sufixos']['DEF1']
        if dados_def1['nome']:
            aplicar_regras_fixas_bloco(equipamento, 'DEF1', dados_def1, SEQ_DEF1_INFO, sobrescrever=False)

        # 3. Preencher EST1 usando a lógica Mista (Saída Formal + Tags Fixas)
        dados_est1 = eq_data['sufixos']['EST1']
        if dados_est1['nome']:
            variaveis_bit = dados_est1['variaveis']
            
            for bit, info in SEQ_EST1_MAPPING.items():
                tag_origem = None
                comment = info.get('comment', '')

                # A. Lógica para Saída Formal do Bloco (lê o XML)
                if info.get('type') == "FORMAL_OUT":
                    out_var = desc_ffb.find(f'./outputVariable[@formalParameter="{info["param"]}"]')
                    tag_origem = out_var.get('effectiveParameter') if out_var is not None and out_var.get('effectiveParameter') else f"{equipamento}_{info['param']}"
                
                # B. Lógica para Tag Fixa (constrói o nome)
                elif info.get('type') == "TAG_FIXA":
                    tag_origem = f"{equipamento}{info['suffix']}"

                # Aplica se for encontrado e o slot estiver vazio
                if tag_origem and not variaveis_bit[bit]['nome']:
                    variaveis_bit[bit]['nome'] = tag_origem
                    variaveis_bit[bit]['comentario'] = comment

    return matriz_equipamentos

def completar_comentarios_faltantes(matriz_equipamentos: Dict[str, Dict[str, Any]], lista_variaveis: List[Dict[str, str]]
) -> Dict[str, Dict[str, Any]]:
    """
    Usa a lista completa de variáveis do UnitPro para preencher os comentários
    dos bits que foram alocados (Tag/Variável) mas ficaram sem comentário.
    """
    print("\nIniciando preenchimento de comentários faltantes...")
    
    # 1. Cria um dicionário de busca rápida (nome_da_tag -> comentario)
    # Isso melhora muito a performance em grandes projetos.
    mapa_comentarios = {
        var['nome']: var['comentario'] 
        for var in lista_variaveis 
        if var['nome']
    }
    
    comentarios_preenchidos = 0
    
    # 2. Itera sobre a matriz de equipamentos
    for equipamento, dados_eq in matriz_equipamentos.items():

        for sufixo, dados_sufixo in dados_eq['sufixos'].items():
            
            # Só processa se a palavra de controle/estado/defeito existir
            if dados_sufixo['nome']:
                variaveis_bit = dados_sufixo['variaveis']
                
                # Itera sobre os 16 bits
                for bit in range(16):
                    bit_info = variaveis_bit[bit]
                    tag_nome = bit_info['nome']
                    comentario_atual = bit_info['comentario']
                    
                    # Condição para preenchimento:
                    # O bit está alocado (tem 'tag_nome') E
                    # O comentário está vazio (ou é um espaço em branco/nulo)
                    if tag_nome and not comentario_atual.strip():
                        
                        # Tenta buscar o comentário no mapa
                        novo_comentario = mapa_comentarios.get(tag_nome)
                        
                        if novo_comentario and novo_comentario.strip():
                            bit_info['comentario'] = novo_comentario.strip()
                            comentarios_preenchidos += 1
    
    print(f"Completado: {comentarios_preenchidos} comentários preenchidos a partir do unitpro.xef.")
    return matriz_equipamentos

# ----------------------------------------------------------------------
# Gera planilha Excel com o resultado final
# ----------------------------------------------------------------------

'''
def ler_titulo_projeto(caminho_arquivo_xef: str) -> str:
    """Lê o atributo 'name' da tag contentHeader no arquivo XEF."""
    try:
        tree = ET.parse(caminho_arquivo_xef)
        root = tree.getroot()
        # Procura por contentHeader no nível raiz
        header = root.find('contentHeader')
        if header is not None:
            return header.get('name', 'Projeto_Sem_Nome')
        return 'Projeto_Sem_Header'
    except (FileNotFoundError, ET.ParseError):
        return 'Projeto_Invalido'
'''

def ler_titulo_projeto(caminho_arquivo_xef: str, lista_variaveis_lidas: List[Dict[str, Any]]) -> str:
    """
    Lê o atributo 'name' da tag contentHeader no arquivo XEF.
    Se o título for "Project", procura por uma tag terminada em '_DCOM' E do tipo 'WORD'
    na lista de variáveis e a utiliza como título.
    """
    
    # --- 1. Lógica Original de Leitura do Título no XML ---
    
    titulo_lido = 'Projeto_Invalido'
    
    try:
        tree = ET.parse(caminho_arquivo_xef)
        root = tree.getroot()
        header = root.find('contentHeader')
        
        if header is not None:
            # Pega o nome. Se não houver, usa 'Projeto_Sem_Nome'.
            titulo_lido = header.get('name', 'Projeto_Sem_Nome')
        else:
            titulo_lido = 'Projeto_Sem_Header'
            
    except (FileNotFoundError, ET.ParseError):
        # Mantém 'Projeto_Invalido'
        pass
        
    # --- 2. Lógica de Verificação e Substituição para "_DCOM" e tipo "WORD" ---
    
    if titulo_lido == "Project":
        
        print("\nAlerta: Título original encontrado é 'Project'. Buscando fallback '_DCOM' (Tipo WORD)...")
        
        # Procura a primeira variável que atenda a ambas as condições
        for variavel in lista_variaveis_lidas:
            nome_variavel = variavel.get('nome', '')
            tipo_variavel = variavel.get('tipo', '')
            
            # Condição A: A tag deve terminar com "_DCOM"
            condicao_dcom = nome_variavel and nome_variavel.endswith('_DCOM')
            
            # Condição B: O tipo deve ser "WORD"
            condicao_word = tipo_variavel == 'WORD'
            
            # Verifica se AMBAS as condições são atendidas
            if condicao_dcom and condicao_word:
                print(f"Substituindo 'Project' pela tag: {nome_variavel}")
                return nome_variavel.removesuffix('_DCOM') # Retorna imediatamente o novo título
                
        # 3. Se o loop terminar sem encontrar a tag "_DCOM" tipo "WORD"
        print("Aviso: Nenhuma tag '_DCOM' do tipo 'WORD' foi localizada na lista de variáveis lidas.")
        return titulo_lido
        
    else:
        # Se o título original for válido e não for "Project", retorna o que foi lido
        return titulo_lido



def desenhar_linhas(ws, equipamentos):
    thin_side = Side(style='thin')
    max_row = (len(equipamentos)+1)*16
    max_col = 49

    for row in range(2, max_row + 1):
        for col in range(1, max_col + 1):
            # Pega bordas existentes para não sobrescrever
            #current_border = ws.cell(row=row, column=col).border

            bottom = None
            # Linha horizontal saltando de 16 em 16
            # Se for linha 3, 19, 35... (start+1)
            if (row - 1) % 16 == 1:  # Ex.: row=3,19,35...
                bottom = thin_side

            # Aplica borda atualizada
            ws.cell(row=row, column=col).border = Border(left=thin_side, right=thin_side, bottom=bottom)


def gerar_planilha_excel_template(matriz_equipamentos: Dict[str, Dict[str, Any]], caminho_unitpro: str, CAMINHO_MODELO: str):


    """
    Gera a planilha de mapeamento de memória a partir de um template, 
    preenchendo os dados da matriz.
    """
    #CAMINHO_MODELO = os.path.join(os.path.dirname(caminho_unitpro), "modelo_mapa_memoria.xlsx")
    
    if not os.path.exists(CAMINHO_MODELO):
        print(f"\nERRO: O arquivo modelo '{CAMINHO_MODELO}' não foi encontrado. Abortando.")
        return

    # 1. Obter Título e Preparar o Nome do Arquivo
    titulo_projeto = ler_titulo_projeto(caminho_unitpro, lista_variaveis_lidas)
    data_atual = datetime.now().strftime("%Y%m%d_%H%M")
    nome_arquivo_destino = f"mapa_de_memoria_{titulo_projeto}_{data_atual}.xlsx"
    caminho_destino = os.path.join(os.path.dirname(caminho_unitpro), nome_arquivo_destino)
    
    # 2. Copiar o arquivo modelo para o destino
    try:
        shutil.copyfile(CAMINHO_MODELO, caminho_destino)
        print(f"\nCopiando template para: {nome_arquivo_destino}")
    except Exception as e:
        print(f"\nERRO ao copiar o template: {e}")
        return

    # 3. Abrir o workbook copiado
    try:
        wb = load_workbook(caminho_destino)
        ws = wb.active
    except Exception as e:
        print(f"\nERRO ao abrir o workbook de destino: {e}")
        return

    # 4. Preencher o Título na célula A1 (mesclada)
    ws['A1'] = titulo_projeto
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws['A1'].font = Font(bold=True, size=14)
    
    # 5. Configuração da Cópia de Linhas (Rows 3 a 18 são o bloco template)
    ROW_START_TEMPLATE = 3
    ROW_END_TEMPLATE = 18
    ROWS_PER_EQUIPMENT = ROW_END_TEMPLATE - ROW_START_TEMPLATE + 1 # 16
    NOME_ABA_MODELO = "Modelo"
    
    # Esta função copiará as células do bloco template para o novo bloco.
    def copy_row_range( target_ws, start_row, end_row, target_start_row):
        source_ws = wb[NOME_ABA_MODELO]  # wb é seu workbook carregad
    
        for i in range(start_row, end_row + 1):
            for col in range(1, 50): # Copia muitas colunas para garantir
                source_cell = source_ws.cell(row=i, column=col)
                target_cell = target_ws.cell(row=target_start_row + (i - start_row), column=col)
                
                # Copia valor e estilo
                target_cell.value = source_cell.value
                if source_cell.has_style:
                    target_cell.font = source_cell.font.copy()
                    target_cell.border = source_cell.border.copy()
                    target_cell.fill = source_cell.fill.copy()
                    target_cell.number_format = source_cell.number_format
                    target_cell.protection = source_cell.protection.copy()
                    target_cell.alignment = source_cell.alignment.copy()
        
        # Copia as mesclagens (Openpyxl torna isso complexo, faremos a mesclagem manual)
        # O template tem mesclagens fixas: Coluna A (Equipamento) e Coluna D (Endereço)
        
        # Coluna A (Equipamento): Mescla de target_start_row até target_start_row + 15
        try:
            target_ws.merge_cells(start_row=target_start_row, start_column=1, 
                                  end_row=target_start_row + ROWS_PER_EQUIPMENT - 1, end_column=1)
        except: pass
        
        # Coluna B (Endereço): Mescla de target_start_row até target_start_row + 15
        # Coluna B e C são Mescladas? O prompt diz A e D. Vamos manter A e B mescladas.
        '''
        try:
             target_ws.merge_cells(start_row=target_start_row, start_column=2, 
                                  end_row=target_start_row + ROWS_PER_EQUIPMENT - 1, end_column=2)
        except: pass
        '''
    # 6. Preenchimento dos Dados
    current_start_row = ROW_START_TEMPLATE
    equipamentos = list(matriz_equipamentos.keys())

    for idx, equipamento in enumerate(equipamentos):
        dados_eq = matriz_equipamentos[equipamento]
       
        # Para o primeiro equipamento (idx=0), usamos o bloco template existente.
        # Para os próximos, copiamos as linhas 3-18 para a posição atual (antes de preencher).
        if idx > 0:
            target_row = current_start_row
            
            # Insere as novas linhas para o template
            ws.insert_rows(target_row, ROWS_PER_EQUIPMENT)
            
            # Copia o template (Rows 3-18) para as novas linhas inseridas
            copy_row_range( ws, ROW_START_TEMPLATE + ROWS_PER_EQUIPMENT, 
                           ROW_END_TEMPLATE + ROWS_PER_EQUIPMENT, target_row)
         
        #código para preencher Coluna B e C (Peso Binário e Índice do Bit)
        start_row_bit_data = current_start_row

        for bit_index in range(ROWS_PER_EQUIPMENT): # 0 a 15
            row_data = start_row_bit_data + bit_index
            
            # 1. Coluna C (Índice do Bit: 0, 1, 2, ...)
            ws.cell(row=row_data, column=3).value = bit_index
            
            # 2. Coluna B (Peso Binário: 1, 2, 4, 8, ...)
            # O valor binário é calculado como 2 elevado à potência do índice do bit (2**bit_index)
            peso_binario = 2 ** bit_index
            ws.cell(row=row_data, column=2).value = peso_binario
            
            # Opcional: Aplicar alinhamento central para ficar organizado
            ws.cell(row=row_data, column=2).alignment = Alignment(horizontal='center', vertical='center')
            ws.cell(row=row_data, column=3).alignment = Alignment(horizontal='center', vertical='center')
        # --- PREENCHIMENTO DO EQUIPAMENTO NA current_start_row ---
          
        # A. Coluna A: Nome do Equipamento (célula mesclada)
        cell_equipamento = ws.cell(row=current_start_row, column=1)
        cell_equipamento.value = equipamento
        # A escrita é transposta
        cell_equipamento.alignment = Alignment(text_rotation=90, horizontal='center', vertical='center')

        # B. Iterar sobre todos os 15 Sufixos na ordem
        col_start_idx = 4 # Coluna D (Address EST1)
        


        for sufixo in SUFIXOS:
            dados_sufixo = dados_eq['sufixos'].get(sufixo, None)
            
            # Se a palavra existe no projeto e foi lida
            if dados_sufixo and dados_sufixo['nome']:
                
                # C1. Endereço da Palavra (Coluna D, G, J, etc. - Mesclada)
                endereco = dados_sufixo.get('endereco', '')

                # Substitui '%MW' por vazio e concatena o '4' no início
                if endereco is not None and '%MW' in endereco:
                    endereco = '4' + endereco.replace('%MW', '') 
                    
                ws.cell(row=current_start_row, column=col_start_idx).value = endereco
                ws.cell(row=current_start_row, column=col_start_idx).alignment = Alignment(text_rotation=90, horizontal='center', vertical='center')
                
                # Mesclagem do endereço (D3:D18, G19:G34, etc.)
                try:
                    ws.merge_cells(start_row=current_start_row, start_column=col_start_idx, 
                                end_row=current_start_row + ROWS_PER_EQUIPMENT - 1, end_column=col_start_idx)
                except: pass

                # C2. Variáveis/Tags e Comentários (Colunas E e F, H e I, etc.)
                variaveis_bit = dados_sufixo['variaveis']
                
                for bit in range(ROWS_PER_EQUIPMENT): # 16 linhas
                    row_offset = bit # 0 a 15
                    row_data = current_start_row + row_offset
                    bit_info = variaveis_bit[bit]
                    
                    # Coluna E/H/K: Tag/Variável (Nome)
                    ws.cell(row=row_data, column=col_start_idx + 1).value = bit_info['nome']
                    
                    # Coluna F/I/L: Comentário
                    ws.cell(row=row_data, column=col_start_idx + 2).value = bit_info['comentario']

            # Move para o próximo bloco de 3 colunas (EST1, EST2, DEF1, etc.)
            col_start_idx += 3 
            
        # D. Move para o início do próximo bloco de linhas (16 linhas abaixo)
                
        current_start_row += ROWS_PER_EQUIPMENT
   

            # 7. Salvar o arquivo
    desenhar_linhas(ws, equipamentos)
    
    try:
        wb.save(caminho_destino)
        print(f"SUCESSO: Planilha formatada salva em '{caminho_destino}'.")
    except Exception as e:
        print(f"\nERRO ao salvar o arquivo Excel: {e}")

# ----------------------------------------------------------------------
# OBSERVAÇÃO DE INTEGRAÇÃO
# ----------------------------------------------------------------------

# O `caminho_unitpro` (caminho_arquivo) deve ser passado para a função para ler o título.
# ----------------------------------------------------------------------
# ORDEM ALFABÉTICA FINAL DA MATRIZ
# ----------------------------------------------------------------------

def ordenar_matriz_alfabeticamente(
    matriz_equipamentos: Dict[str, Dict[str, Any]], habilitar_ordenacao: bool = True) -> Dict[str, Dict[str, Any]]:

    if not habilitar_ordenacao:
        print("Ordenação alfabética desabilitada. Retornando matriz original.")
        return matriz_equipamentos
    
    print("Ordenando matriz de equipamentos alfabeticamente...")
    
    # Obtém as chaves (nomes dos equipamentos) e as ordena
    chaves_ordenadas = sorted(matriz_equipamentos.keys())
    
    # Cria um novo dicionário (que mantém a ordem de inserção)
    matriz_ordenada = {}
    
    for chave in chaves_ordenadas:
        matriz_ordenada[chave] = matriz_equipamentos[chave]
        
    return matriz_ordenada

# ----------------------------------------------------------------------
# EXECUÇÃO PRINCIPAL
# ----------------------------------------------------------------------

if __name__ == "__main__":

    # --- DEFINIÇÃO UNIVERSAL DO CAMINHO BASE ---
    # Essa lógica funciona tanto para o script .py quanto para o executável .exe (frozen)
    if getattr(sys, 'frozen', False):
        # Se estiver rodando como executável (PyInstaller), usa o caminho do binário.
        diretorio_script = os.path.dirname(sys.executable)
    else:
        # Se estiver rodando como script Python (.py), usa o caminho do arquivo de script.
        # É fundamental usar o try-except ou um método robusto para evitar erros ao ser chamado de outro diretório.
        try:
            diretorio_script = os.path.dirname(os.path.abspath(__file__))
        except NameError:
            # Fallback caso __file__ não esteja definido (raro, mas seguro)
            diretorio_script = os.path.getcwd() 

    # --- Configuração de Caminhos ---
    caminho_unitpro = os.path.join(diretorio_script, ARQUIVO_UNITPRO)
    CAMINHO_DO_TEMPLATE_EXCEL = os.path.join(diretorio_script, "modelo_mapa_memoria.xlsx")


    #diretorio_script = os.path.dirname(os.path.abspath(__file__))
    #caminho_unitpro = os.path.join(diretorio_script, ARQUIVO_UNITPRO)
    #CAMINHO_DO_TEMPLATE_EXCEL = os.path.join(diretorio_script, "modelo_mapa_memoria.xlsx")

    print(f"Iniciando processamento do arquivo: {caminho_unitpro}")
    
    # 1. Leitura e Catalogação das Variáveis
    lista_variaveis_lidas = ler_variaveis_unitpro(caminho_unitpro)
    matriz_equipamentos = catalogar_variaveis(lista_variaveis_lidas)
    
    # --- FASE DE PREENCHIMENTO BASEADO EM BLOCOS DE FUNÇÃO (Prioridade 1 - Sobrescreve) ---
    
    # 2. Bloco MOTOR (DEFx, CMDx, ESTx) - PRIORIDADE 1
    print("\nProcurando blocos MOTOR[...] para preencher as palavras de controle e defeito (PRIORIDADE 1)...")
    matriz_equipamentos = preencher_bits_bloco_motor(caminho_unitpro, matriz_equipamentos)

    # 3. Bloco VALVULA (DEFx, CMDx, ESTx) - PRIORIDADE 1
    print("Procurando blocos VAL[...] para preencher as palavras de controle e defeito (PRIORIDADE 1)...")
    matriz_equipamentos = preencher_bits_bloco_valvula(caminho_unitpro, matriz_equipamentos)
    
    # --- FASE DE PREENCHIMENTO BASEADO EM CONEXÕES XML (Prioridade 2/3/4 - Não Sobrescreve) ---
    
    # 4. Bloco BIT_TO_WORD - PRIORIDADE 2
    print("Procurando blocos BIT_TO_WORD para preencher os bits restantes (PRIORIDADE 2)...")
    matriz_equipamentos = preencher_bits_bit_to_word(caminho_unitpro, matriz_equipamentos)
    
    # 5. Bloco SET_BIT - PRIORIDADE 3
    print("Procurando blocos SET_BIT para preencher bits individuais (PRIORIDADE 3)...")
    matriz_equipamentos = preencher_bits_set_bit(caminho_unitpro, matriz_equipamentos)
    
    # 6. Bloco SEQ - PRIORIDADE 4
    print("Procurando blocos SEQ para preencher EST1, DEF1 e CMD1 (PRIORIDADE 4)...")
    matriz_equipamentos = preencher_bits_bloco_seq(caminho_unitpro, matriz_equipamentos)

    # 7. NOVO PASSO: Complementar Comentários Faltantes (Usa a lista original)
    matriz_final = completar_comentarios_faltantes(matriz_equipamentos,lista_variaveis_lidas)
    
    # 8. Ordenação Alfabética Final da Matriz
  
    matriz_final = ordenar_matriz_alfabeticamente(matriz_final, Habilitar_ordenacao)


    # 9. Geração da Planilha Excel (Usa a matriz com comentários complementados)
    gerar_planilha_excel_template(matriz_final,caminho_unitpro,CAMINHO_DO_TEMPLATE_EXCEL)
    # ... (Restante da Demonstração/Geração de Documentação)
    print("\n--- Processamento Concluído ---")